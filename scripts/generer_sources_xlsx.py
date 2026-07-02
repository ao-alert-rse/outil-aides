"""
Génère SOURCES.xlsx — fichier consolidé de toutes les sources de l'outil.
4 onglets : Aides, OPCO, APIs & Outils, Procédure MAJ.

Usage : python scripts/generer_sources_xlsx.py
À régénérer après chaque modification de la BDD AIDES.
"""
import re, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
HTML = ROOT / "outil-aides-rse.html"
OUT = ROOT / "SOURCES.xlsx"

# ─── Styles ───
HEADER_FILL = PatternFill("solid", fgColor="2C5F2D")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 30


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─── Onglet 1 : Aides (extrait depuis HTML) ───
def parse_aides_from_html() -> list[dict]:
    html = HTML.read_text(encoding="utf-8")
    # Capture le bloc const AIDES = [ ... ];
    m = re.search(r"const AIDES = \[(.*?)\n\];", html, re.DOTALL)
    if not m:
        return []
    body = m.group(1)
    aides = []
    # Découpe en blocs { ... } séparés par des virgules au niveau racine
    # Approche simple : split sur "  {" en début de ligne
    blocks = re.split(r"\n\s{2}\{\n", body)
    for blk in blocks[1:]:
        d = {}
        for field in ("id", "type", "nom", "description", "taux", "plafond", "origine", "lien", "derniere_verif", "region"):
            mm = re.search(rf'{field}:\s*"([^"]*)"', blk)
            if mm:
                d[field] = mm.group(1)
        for field in ("effectif_min", "effectif_max"):
            mm = re.search(rf"{field}:\s*(\d+|null)", blk)
            if mm:
                d[field] = mm.group(1)
        # opco: ["a","b"]
        mm = re.search(r'opco:\s*\[([^\]]*)\]', blk)
        if mm:
            d["opco"] = ", ".join(re.findall(r'"([^"]+)"', mm.group(1)))
        if d.get("id"):
            aides.append(d)
    return aides


def sheet_aides(wb):
    ws = wb.create_sheet("Aides", 0)
    headers = ["ID", "Type", "Nom", "Origine", "OPCO ciblé", "Effectif min", "Effectif max",
               "Région", "Taux", "Plafond", "Lien officiel", "Dernière vérif", "Description"]
    ws.append(headers)
    for a in parse_aides_from_html():
        ws.append([
            a.get("id", ""), a.get("type", ""), a.get("nom", ""), a.get("origine", ""),
            a.get("opco", ""), a.get("effectif_min", ""), a.get("effectif_max", ""),
            a.get("region", ""), a.get("taux", ""), a.get("plafond", ""),
            a.get("lien", ""), a.get("derniere_verif", ""), a.get("description", ""),
        ])
    style_header(ws, len(headers))
    autosize(ws, [22, 8, 45, 22, 18, 12, 12, 8, 30, 30, 50, 14, 70])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP


# ─── Onglet 2 : OPCO ───
OPCO_DATA = [
    ("atlas", "OPCO Atlas", "Banque, assurance, finance, BET, conseil", "https://www.opco-atlas.fr/criteres-financement/", "Couvert (BET)", "Plafonds détaillés par tranche d'effectif"),
    ("opco2i", "OPCO 2i", "Industrie (métallurgie, chimie, plasturgie, textile…)", "https://www.opco2i.fr/formation-et-financement/alternance-et-apprentissage/", "Couvert partiellement (alternance OK, PDC = espace connecté)", "Chiffres alternance vérifiés via PDFs officiels"),
    ("opcoep", "OPCO EP", "Entreprises de proximité (coiffure, fleuristes, boucherie, etc.)", "https://www.opcoep.fr/criteres-de-financement", "Couvert (Coiffure)", "Détails branche par branche"),
    ("afdas", "AFDAS", "Culture, médias, audiovisuel, sport, loisirs, communication", "https://www.afdas.com/actualites/entreprise/criteres-de-financements-2026-ce-qui-change-pour-les-entreprises.html", "Couvert (PDC + alternance)", "Critères 2026 : 1 100 €/an (<11) et 1 800 €/an (11-49) — 40 €/h"),
    ("akto", "AKTO", "Services à forte intensité de main-d'œuvre (propreté, sécurité, restauration collective…)", "https://www.akto.fr/financer-une-formation/regles-de-prise-en-charge/", "Couvert partiellement (Propreté IDCC 3043)", "PDF officiel branche par branche — voir data/akto_*.pdf"),
    ("constructys", "Constructys", "Bâtiment, travaux publics, négoce de matériaux", "https://www.constructys.fr/decouvrez-les-modalites-de-participation-financiere-2026/", "Couvert (PDC Bâtiment + TP + alternance)", "Plafonds 2026 : 10-32 €/h selon segment, max 8 000 €/an pour TP avec accord"),
    ("ocapiat", "OCAPIAT", "Agriculture, pêche, agroalimentaire, coopératives", "https://www.ocapiat.fr/wp-content/uploads/Regles-de-prise-en-charge-2026-OCAPIAT.pdf", "Couvert (PDC + bilan + Boost)", "PDC 2026 : 4 500 €/an (<11) et 9 000 €/an (11-49) — chiffres à reconfirmer (site 503 le 25/06)"),
    ("opco-mobilites", "OPCO Mobilités", "Transport routier, ferroviaire, fluvial, urbain, services automobile", "https://www.opcomobilites.fr/media/2026/03/Services-de-lautomobile_24032026-1.pdf", "Couvert partiellement (Services automobile)", "Plafonds 2026 par tranche : 1500/1800/2100/2400/2700 € — autres branches à enrichir"),
    ("sante", "Opco Santé", "Santé, médico-social privé (SSSMS, HP, SPSTI, hors CC)", "https://www.opco-sante.fr/le-plan-de-developpement-des-competences-pdc", "Couvert partiellement (4 branches + apprenti TH)", "⚠ Synthèses chiffrées par branche protégées — chiffres à compléter via espace OPCO Santé"),
    ("cohesion-sociale", "Uniformation / OPCO Cohésion Sociale", "Économie sociale, habitat social, secteur associatif, insertion (SIAE)", "https://www.uniformation.fr/entreprise/financements/frais-annexes-et-couts-pedagogiques", "Couvert (PDC + SIAE + AEFMA + alternance)", "PDC <50 : 5 000 €/an + 2 000 €/jour (présentiel) — AEFMA 230-345 €/mois"),
    ("commerce", "OPCO Commerce", "Commerce (grande distribution, succursalistes, grossistes, commerce de détail)", "https://www.lopcommerce.com/media/bqymkcgh/cpc_spp_commerce.pdf", "Couvert (PDC + alternance + Click&Form)", "PDC TPE 1 000 €, PME 1 500 €/an + coût-contrat 5 470 à 7 000 € selon niveau RNCP"),
]


def sheet_opco(wb):
    ws = wb.create_sheet("OPCO")
    headers = ["Slug", "Nom officiel", "Secteurs couverts", "URL critères de financement", "Statut BDD", "Notes"]
    ws.append(headers)
    for row in OPCO_DATA:
        ws.append(list(row))
    style_header(ws, len(headers))
    autosize(ws, [18, 35, 50, 55, 22, 40])
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP


# ─── Onglet 3 : APIs & Outils ───
APIS = [
    ("API recherche-entreprises", "DINUM / Etalab", "SIRET → NAF, effectif, région, raison sociale, liste_idcc", "https://recherche-entreprises.api.gouv.fr/", "Gratuit, sans clé, CORS ouvert", "En production"),
    ("Table IDCC → OPCO", "Ministère du Travail", "Mapping 332 IDCC vers 11 OPCO", "https://bretagne.dreets.gouv.fr/sites/bretagne.dreets.gouv.fr/IMG/xlsx/table_idcc-opco-19072019-ministere_du_travail_.xlsx", "Mirror DREETS Bretagne (l'URL officielle Ministère 404)", "Date fichier : 19/07/2019"),
    ("Table SIRET-OPCO « SIRO »", "France Compétences", "SIRET → IDCC → OPCO (mensuel)", "https://www.data.gouv.fr/datasets/table-siret-opco", "CSV 101 Mo — trop lourd pour embarquer", "Pour vérif ponctuelle"),
    ("Quel-est-mon-OPCO", "France Compétences", "Identification interactive d'OPCO", "https://quel-est-mon-opco.francecompetences.fr/", "Pas d'API publique", "Backup manuel si auto-détection échoue"),
    ("Plateforme Agir pour la Transition", "ADEME", "Catalogue exhaustif des aides ADEME entreprises", "https://agirpourlatransition.ademe.fr/entreprises/aides-financieres", "Portail à consulter mensuellement", "Source #1 pour aides TEE"),
    ("Mission Transition Écologique", "beta.gouv.fr", "Guichet unique TPE/PME pour aides transition", "https://mission-transition-ecologique.beta.gouv.fr/", "Portail filtre par taille/secteur", "Recommandé TPE/PME"),
    ("Calculateur CEE", "ADEME", "Simulation valorisation CEE par projet", "https://calculateur-cee.ademe.fr/", "Outil de simulation", "Pour estimer un projet avant montage"),
    ("Annuaire ADEME Régions", "ADEME", "Représentants ADEME en région", "https://www.ademe.fr/lademe/notre-organisation/implantations/", "Annuaire des directions régionales", "Contacts directs par région"),
    ("Codes NAF officiels", "INSEE", "Nomenclature d'activités française (NAF rév. 2)", "https://www.insee.fr/fr/information/2120875", "Référence", "Pour comprendre le code NAF d'une entreprise"),
    ("Tranches d'effectif INSEE", "INSEE", "Codes de tranches d'effectif salarié", "https://www.insee.fr/fr/information/2028155", "Référence", "Borne haute utilisée par le matcher"),
]


def sheet_apis(wb):
    ws = wb.create_sheet("APIs & Outils")
    headers = ["Nom", "Fournisseur", "Description", "URL", "Statut / contraintes", "Notes"]
    ws.append(headers)
    for row in APIS:
        ws.append(list(row))
    style_header(ws, len(headers))
    autosize(ws, [30, 22, 50, 55, 35, 40])
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP


# ─── Onglet 4 : Procédure MAJ ───
PROC = [
    ("1", "Vérifier les liens", "Lance python scripts/verifier_liens.py", "Mensuelle", "2 min"),
    ("2", "Table IDCC → OPCO", "Vérifier si MAJ Ministère ; sinon python scripts/maj_idcc_opco.py", "Trimestrielle", "5 min si MAJ"),
    ("3", "Plafonds OPCO Atlas", "Comparer avec opco-atlas.fr/criteres-financement/", "Annuelle", "5 min"),
    ("4", "Plafonds OPCO EP", "Comparer avec opcoep.fr/criteres-de-financement", "Annuelle", "5 min"),
    ("5", "Plafonds OPCO 2i", "À enrichir — page JS dynamique", "À couvrir", "Session dédiée"),
    ("6", "Plafonds 9 autres OPCO", "À enrichir — voir onglet OPCO", "À couvrir", "Session dédiée"),
    ("7", "Aides ADEME / Agir pour la Transition", "Vérifier nouveautés sur le portail", "Semestrielle", "10 min"),
    ("8", "Aides Bpifrance / France 2030", "Vérifier bpifrance.fr/taxonomy/term/1076 + AAP France 2030", "Trimestrielle", "10 min"),
    ("9", "Aides CEE", "Vérifier mises à jour ecologie.gouv.fr", "Annuelle", "5 min"),
    ("10", "FEDER / Innovation Fund UE", "Vérifier nouveaux appels à projets sur europe-en-france.gouv.fr et cinea.ec.europa.eu", "Semestrielle", "10 min"),
    ("11", "PCRH / Travail Emploi", "Vérifier travail-emploi.gouv.fr/la-prestation-de-conseil-en-ressources-humaines-pour-les-tpe-pme", "Annuelle", "5 min"),
    ("12", "Aides Région IdF (iledefrance.fr)", "Vérifier iledefrance.fr/aides-et-appels-a-projets (hors CCI — recentrage V3)", "Semestrielle", "5 min"),
]


def sheet_procedure(wb):
    ws = wb.create_sheet("Procédure MAJ")
    headers = ["#", "Étape", "Action", "Fréquence", "Effort"]
    ws.append(headers)
    for row in PROC:
        ws.append(list(row))
    style_header(ws, len(headers))
    autosize(ws, [5, 32, 60, 18, 18])
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP

    # Note finale
    last = ws.max_row + 2
    ws.cell(row=last, column=1, value="Règle d'or :").font = Font(bold=True, color="C00000")
    ws.cell(row=last, column=2, value="Aucun chiffre inventé. Si non vérifiable sur source officielle → \"N/A — voir lien\".")
    ws.merge_cells(start_row=last, start_column=2, end_row=last, end_column=5)
    ws.cell(row=last, column=2).alignment = WRAP


# ─── MAIN ───
def main():
    wb = Workbook()
    # supprime la feuille par défaut
    del wb["Sheet"]
    sheet_aides(wb)
    sheet_opco(wb)
    sheet_apis(wb)
    sheet_procedure(wb)
    wb.save(OUT)
    print(f"OK — {OUT.name} ({OUT.stat().st_size} octets)")


if __name__ == "__main__":
    main()
